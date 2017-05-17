#!/usr/bin/env python
# -*- coding: utf-8 -*-

######################################################
#
# File Name:  data_to_excel_by_sheet.py
#
# Function:   
#
# Usage:  
#
# Input:  
#
# Output:	
#
# Author: panwenhai
#
# Create Time:    2017-05-16 19:16:43
#
######################################################

import sys
reload(sys)
sys.setdefaultencoding("utf-8")
import os
import time
from datetime import datetime, timedelta

import xlrd
import xlwt
# workbook相关
from openpyxl.workbook import Workbook
# ExcelWriter，封装了很强大的excel写的功能
from openpyxl.writer.excel import ExcelWriter
# 一个eggache的数字转为列字母的方法
from openpyxl.utils import get_column_letter
from openpyxl.reader.excel import load_workbook


def main():

    src_file = sys.argv[1]
    dst_file = sys.argv[2]

    src = open(src_file, "r")

    wb = Workbook()

    grade_map = {}

    head_line = []
    head_line.append("学生id")
    head_line.append("学生姓名")
    head_line.append("城市")
    head_line.append("科目")
    head_line.append("年级")
    head_line.append("班型")
    head_line.append("课次")
    head_line.append("预习")
    head_line.append("课前测")
    head_line.append("课后测")
    head_line.append("作业")
    head_line.append("复习巩固")

    i = 0

    for line in src:
        i += 1
        if i % 10000 == 0:
            print i

        ary = line.decode("utf-8").strip().split(",")

        data = [a[1:-1] for a in ary]
        grade = data[4]

        if grade not in grade_map:

            grade_map[grade] = [wb.create_sheet(title=grade), 1]

            for col in range(1, len(head_line) + 1):
                grade_map[grade][0].cell(column=col, row=1, value = "%s" % (head_line[col-1]))

            grade_map[grade][1] += 1

        for col in range(1, len(data) + 1):
            grade_map[grade][0].cell(column=col, row=grade_map[grade][1], value="%s" % (data[col-1]))

        grade_map[grade][1] += 1

    wb.save(filename=dst_file)

    return


if __name__ == "__main__":

    main()



